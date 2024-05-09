import meraki
from openpyxl import load_workbook
import csv
import os
import traceback


# Dashboard API key
api_key = "your_org_api_key"
dashboard = meraki.DashboardAPI(api_key, caller="PyCharm/2023.3 JetBrains")

# Get the organizations that the user has privileges on
org = dashboard.organizations.getOrganizations()
org_id = org[0]['id']


# Make a dictionary of the file content
def file_to_dict(file):
    _, file_extension = os.path.splitext(file)   # Get the file extension
    if file_extension.lower() == '.csv':   # Is file extension csv or xlsx
        return csv_to_dict(file)
    elif file_extension.lower() == '.xlsx':
        return xlsx_to_dict(file)
    else:
        return 'Unknown'


# Create a dictionary of a csv file
def csv_to_dict(file):
    data = []   # Initialize an empty list to store rows
    with open(file, newline='') as f:
        reader = csv.reader(f, delimiter=',')
        headers = next(reader)  # Read the headers
        for row in reader:
            row_dict = dict(zip(headers, row))  # Create a dictionary for each row
            data.append(row_dict)
    return data


# Create a dictionary of a xlsx file
def xlsx_to_dict(file):
    wb = load_workbook(file, data_only=True)  # Load the workbook
    ws = wb.active   # Select the active worksheet
    data = []     # Initialize an empty list to store rows
    headers = None
    for row in ws.iter_rows(values_only=True):   # Iterate over rows in the worksheet
        if headers is None:   # Assuming the first row contains headers
            headers = row
        else:
            data.append(dict(zip(headers, row)))   # Create a dictionary mapping headers to row values
    return data


# Retrieve network ids
def get_network_ids():
    nets = {}
    for network in dashboard.organizations.getOrganizationNetworks(org_id):
        nets[network['name']] = network['id']
    return nets


# Create a network in an organization
def create_network(file):
    net_dict = file_to_dict(file)
    networks = []
    for network in net_dict:
        networks.append(
            dashboard.organizations.createOrganizationNetwork(
                org_id,
                # copyFromNetworkId=network['copy_from_net_id'],
                name=network['name'],
                productTypes=[product.strip() for product in network['product_types'].split(",")],
                tags=[tag.strip() for tag in network['tags'].split(",")],
                timeZone=network['time_zone'],
                notes=network['notes']
            )
        )
    return networks


# Claim device(s) into a network
def claim_device(net_id, file):
    device_dict = file_to_dict(file)
    serials = []
    for item in device_dict:
        serials.append(item['serial'])
    return (
        dashboard.networks.claimNetworkDevices(
            net_id,
            serials
        )
    )


# Create a new dashboard administrator
def create_admin(file):
    admin_dict = file_to_dict(file)
    admins = []
    for admin in admin_dict:
        if admin['organization access'] in ['enterprise', 'read-only', 'full']:
            admins.append(
                dashboard.organizations.createOrganizationAdmin(
                    org_id,
                    name=admin['name'],
                    orgAccess=admin['organization access'],
                    authenticationMethod=admin['authentication method'],
                    email=admin['email'],
                    tags=[{'tag': tag.strip('()').split(':')[0],
                           'access': tag.strip('()').split(':')[1]}
                          for tag in admin['(tag:access)'].split(",")]
                )
            )
        elif admin['organization access'] == 'none':
            admins.append(
                dashboard.organizations.createOrganizationAdmin(
                    org_id,
                    name=admin['name'],
                    orgAccess= 'none',
                    authenticationMethod=admin['authentication method'],
                    email=admin['email'],
                    networks=[{'id': networks[net_str.strip('()').split(':')[0]],
                               'access': net_str.strip('()').split(':')[1]}
                              for net_str in admin['(network:access)'].split(",")],
                    tags=[{'tag': tag.strip('()').split(':')[0],
                           'access': tag.strip('()').split(':')[1]}
                          for tag in admin['(tag:access)'].split(",")]
                )
            )
    return admins


# Get vlan cidr
def get_vlan_cidr(net_id, vlan):
    for vl in dashboard.appliance.getNetworkApplianceVlans(net_id):
        if vl['name'].lower() == vlan.lower():
            return vl['subnet']
    return vlan


# Enable vlans for a specific network
def enable_vlan_network(net_id):
    dashboard.appliance.updateNetworkApplianceVlansSettings(
        net_id,
        vlansEnabled=True
    )


# Create vlans for a network
def create_vlan(net_id, file):
    enable_vlan_network(net_id)
    vlan_dict = file_to_dict(file)
    vlans = []
    for row in vlan_dict:
        vlans.append(
            dashboard.appliance.createNetworkApplianceVlan(
                net_id,
                id=row["id"],
                name=row["name"],
                subnet=row["subnet"],
                applianceIp=row["interface_ip"],
                # groupPolicyId=row["group_policy_id"],
                templateVlanType=row["template"],
                cidr=row['temp_cidr'],
                mask=row['temp_mask'],
                # ipv6={'enabled': True, 'prefixAssignments': [
                #    {'autonomous': False, 'staticPrefix': '2001:db8:3c4d:15::/64',
                #     'staticApplianceIp6': '2001:db8:3c4d:15::1',
                #     'origin': {'type': 'internet', 'interfaces': ['wan0']}}]},
                mandatoryDhcp={'enabled': row["mandatory_dhcp"]}
            )
        )
    return vlans


# Configure management interface(s)
def mngt_interface(file):
    mngt_int_dict = file_to_dict(file)
    mn_inter = []
    for interface in mngt_int_dict:
        wan = {'wanEnabled': interface['wan_enabled'],
               'usingStaticIp': interface['using_static_ip'],
               'staticIp': interface['ip'],
               'staticSubnetMask': interface['subnet_mask'],
               'staticGatewayIp': interface['gateway_ip'],
               'staticDns': [dns.strip() for dns in interface['static_dns'].split(",")],
               'vlan': interface['vlan']}
        if interface['wan'] == 'wan1':
            mn_inter.append(
                dashboard.devices.updateDeviceManagementInterface(
                    serial=interface['serial'],
                    wan1=wan
                )
            )
        elif interface['wan'] == 'wan2':
            mn_inter.append(
                dashboard.devices.updateDeviceManagementInterface(
                    serial=interface['serial'],
                    wan2=wan,
                )
            )
    return mn_inter


# Update device settings
def update_device(file):
    upd_dev_dict = file_to_dict(file)
    devices = []
    for device in upd_dev_dict:
        devices.append(
            dashboard.devices.updateDevice(
                serial=device['serial'],
                name=device['name'],
                tags=[tag.strip() for tag in device['tags'].split(",")],
                notes=device['notes'],
                # address='',
                # lat=37.4180951010362,
                # lng=-122.098531723022,
                # floorPlanId='',
                # switchProfileId='',
                moveMapMarker=device['moveMapMarker']
            )
        )
    return devices


# Configure switch ports
def update_sw_ports(file):
    sw_port_dict = file_to_dict(file)
    conf_ports = []
    for port in sw_port_dict:
        conf_ports.append(
            dashboard.switch.updateDeviceSwitchPort(
                serial=port['serial'],
                portId=port['interface_id'],
                name=port['name'],
                tags=[tag.strip() for tag in port['tags'].split(",")],
                enabled=port['enabled'],
                # poeEnabled=True,
                type=port['type'],
                vlan=port['vlan'],
                # voiceVlan=None,
                # allowedVlans='all',
                # isolationEnabled=False,
                # rstpEnabled=True,
                # stpGuard='disabled',
                linkNegotiation=port['link_nego'],
                # portScheduleId=None,
                # udld='Alert only',
                # accessPolicyType='Open',
                # daiTrusted=False,
                # profile={'enabled': False, 'id': '', 'iname': None}
            )
        )
    return conf_ports


# Update the attributes of an MR SSID
def config_ssid(net_id, file):
    ssid_dict = file_to_dict(file)
    conf_ssid = []
    for row in ssid_dict:
        conf_ssid.append(
            dashboard.wireless.updateNetworkWirelessSsid(
                net_id,
                number=row['number'],
                name=row['name'],
                enabled=row['enabled'],
                authMode=row['auth_mode'],
                # nterpriseAdminAccess='access enabled',
                encryptionMode=row['encrypt_mode'],
                psk=row['password'],
                wpaEncryptionMode=row['wpa_encrypt_mode'],
                # dot11w={'enabled': True, 'required': False},
                # dot11r={'enabled': True, 'adaptive': True},
                splashPage=row['splash_mode'],
                # splashGuestSponsorDomains=['example.com'],
                # oauth={'allowedDomains': ['example.com']},
                # localRadius={'cacheTimeout': 60, 'passwordAuthentication': {'enabled': False},
                #              'certificateAuthentication': {'enabled': True, 'useLdap': False, 'useOcsp': True,
                #                                            'ocspResponderUrl': 'http://ocsp-server.example.com',
                #                                            'clientRootCaCertificate': {
                #                                                'contents': ''}}},
                # ldap={'servers': [{'host': '127.0.0.1', 'port': 389}],
                #       'credentials': {'distinguishedName': 'cn=user,dc=example,dc=com', 'password': 'password'},
                #       'baseDistinguishedName': 'dc=example,dc=com', 'serverCaCertificate': {
                #         'contents': ''}},
                # activeDirectory={'servers': [{'host': '127.0.0.1', 'port': 3268}],
                #                  'credentials': {'logonName': 'user', 'password': 'password'}},
                # radiusServers=[{'host': '0.0.0.0', 'port': 3000, 'secret': 'secret-string', 'radsecEnabled': True,
                #                 'openRoamingCertificateId': 2,
                #                 'caCertificate': ''}],
                # radiusProxyEnabled=False,
                # radiusTestingEnabled=True,
                # radiusCalledStationId='00-11-22-33-44-55:AP1',
                # radiusAuthenticationNasId='00-11-22-33-44-55:AP1',
                # radiusServerTimeout=5,
                # radiusServerAttemptsLimit=5,
                # radiusFallbackEnabled=True,
                # radiusCoaEnabled=True,
                # radiusFailoverPolicy='Deny access',
                # radiusLoadBalancingPolicy='Round robin',
                # radiusAccountingEnabled=True,
                # radiusAccountingServers=[
                #     {'host': '0.0.0.0', 'port': 3000, 'secret': 'secret-string', 'radsecEnabled': True,
                #      'caCertificate': ''}],
                # radiusAccountingInterimInterval=5,
                # radiusAttributeForGroupPolicies='Filter-Id',
                ipAssignmentMode=row['ip_assign_mode'],
                useVlanTagging=row['vlan_tagging'],
                # concentratorNetworkId='N_24329156',
                # secondaryConcentratorNetworkId='disabled',
                # disassociateClientsOnVpnFailover=False,
                vlanId=row['tagged_vlan'],
                defaultVlanId=row['tagged_vlan'],
                # apTagsAndVlanIds=[{'tags': ['tag1', 'tag2'], 'vlanId': 100}],
                # walledGardenEnabled=True,
                # walledGardenRanges=['example.com', '1.1.1.1/32'],
                # gre={'concentrator': {'host': '192.168.1.1'}, 'key': 5},
                # radiusOverride=False,
                # radiusGuestVlanEnabled=True,
                # radiusGuestVlanId=1,
                # minBitrate=5.5,
                # bandSelection='5 GHz band only',
                # perClientBandwidthLimitUp=0,
                # perClientBandwidthLimitDown=0,
                # perSsidBandwidthLimitUp=0,
                # perSsidBandwidthLimitDown=0,
                # lanIsolationEnabled=True,
                visible=row['visable'],
                # availableOnAllAps=False,
                # availabilityTags=['tag1', 'tag2'],
                # mandatoryDhcpEnabled=False,
                # adultContentFilteringEnabled=False,
                # dnsRewrite={'enabled': True, 'dnsCustomNameservers': ['8.8.8.8', '8.8.4.4']},
                # speedBurst={'enabled': True},
                # namedVlans={'tagging': {'enabled': True, 'defaultVlanName': 'My VLAN',
                #                         'byApTags': [{'tags': ['tag1', 'tag2'], 'vlanName': 'My VLAN'}]},
                #             'radius': {'guestVlan': {'enabled': True, 'name': 'Guest VLAN'}}}
            )
        )
    return conf_ssid


# Update the L3 outbound firewall rules of an MX network
def config_l3_out_acl(net_id, file):
    l3_acl_dict = file_to_dict(file)
    rules = []
    for acl in l3_acl_dict:
        rules.append(acl)
    conf_acl = (
        dashboard.appliance.updateNetworkApplianceFirewallL3FirewallRules(
            networkId=net_id,
            rules=rules
        )
    )
    return conf_acl


if __name__ == "__main__":
    try:
        create_network('network.csv')
        networks = get_network_ids()
        elfa_id = networks['Elfa-exjobb']
        claim_device(elfa_id,'deviceSettings.csv')
        create_admin('admins.csv')
        create_vlan(elfa_id,'vlans.csv')
        update_device('deviceSettings.csv')
        update_sw_ports('swPorts.csv')
        mngt_interface('mngtInterface.csv')
        config_ssid(elfa_id, 'ssids.csv')
        config_l3_out_acl(elfa_id, 'L3outACL.csv')
        print('-' * 120)
        print("!!!THE MISSION IS ACCOMPLISHED!!!".center(120))
    except (TypeError, ValueError, KeyError, AssertionError, IndexError, FileNotFoundError) as e:
        print('-' * 120)
        print("!!!Something went WRONG!!!".center(120))
        print('-' * 120)
        print("Error:", str(e))
        traceback.print_exc()
    except Exception as e:
        print('-' * 120)
        print("!!!Something went WRONG!!!".center(120))
        print('-' * 120)
        print(f"Level:     {e.tag}\n"
              f"Operation: {e.operation}\n"
              f"Status:    {e.status} ({e.reason})\n"
              f"Errors:")
        counter = 1
        for error in e.message['errors']:
            print(f"       {counter}. {error}")
            counter += 1
    print('-' * 120)
